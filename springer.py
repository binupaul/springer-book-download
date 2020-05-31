#!/usr/bin/env python3

import re
import os
import sys
from urllib.parse import urljoin
from http.cookiejar import MozillaCookieJar
import argparse
import openpyxl
import requests
from bs4 import BeautifulSoup


def run():
    parser = argparse.ArgumentParser(description="Download books from Springer")
    parser.add_argument("xlfile", help="the Excel file containing the book list")
    parser.add_argument("cookies", help="the exported cookies file")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--list", "-l", help="only list books", action="store_true")
    group.add_argument("--topics", "-a", help="only list topics", action="store_true")
    parser.add_argument("--interactive", "-p", help="prompt for user input if required",
                        action="store_true")
    parser.add_argument("--topic", "-t", help="use the given topic")
    parser.add_argument("--isbn", "-i", help="use the given ISBN")
    args = parser.parse_args()

    try:
        work_book = openpyxl.load_workbook(args.xlfile)
    except OSError as err:
        print("Error:", err)
        return

    work_sheet = work_book.active

    if args.list:
        books = do_list_books(args, work_sheet)
        print_books(books)
    elif args.topics:
        topics = do_list_topics(args, work_sheet)
        print_topics(topics)
    else:
        books = do_list_books(args, work_sheet)
        if args.interactive:
            filtered = []
            for book in books:
                print(book[0])
                if confirm_yes_no("Add to download list [yes]? "):
                    filtered.append(book)
            books = filtered
        download_books(books, args.cookies)

def download_books(books, cookiesfile):
    cookiejar = MozillaCookieJar(cookiesfile)
    cookiejar.load(ignore_discard=True)

    for book in books:
        print(f"Attempting to download book {book[0]}")
        response = requests.get(book[4], cookies=cookiejar)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            anchor = soup.find("a", title="Download this book in PDF format")
            if not anchor:
                print(f"Unable to download {book[0]}")
                continue
            final_url = urljoin(book[4], anchor.get('href'))
            download_book(final_url, book[0], book[3], cookiejar)
        else:
            print(f"Unable to download {book[0]}")

def download_book(url, title, topic, cookiesjar):
    response = requests.get(url, cookies=cookiesjar)
    if response.status_code == 200:
        fname = get_download_filename(response)
        complete_name = os.path.join(topic, fname)
        if not os.path.exists(topic):
            os.mkdir(topic)
        with open(complete_name, "wb") as file_handle:
            file_handle.write(response.content)
            print(f"Saved {title} to {fname} in directory {topic}")
    else:
        print(f"Unable to download {title}")

def get_download_filename(response):
    header = response.headers['content-disposition']
    fname = re.findall("filename=(.+)", header)[0]
    return fname

def print_topics(topics):
    for topic in topics:
        print(topic)

def print_books(books):
    for book in books:
        print('\n'.join(book), '\n')

def do_list_books(args, work_sheet):
    books = []
    for row in work_sheet.iter_rows():
        to_add = True
        for cell in row:
            if cell.row == 1:
                to_add = False
                break
            if cell.column == 'L':
                topic = cell.value
            elif cell.column == 'G':
                isbn = cell.value
            elif cell.column == 'A':
                title = cell.value
            elif cell.column == 'B':
                author = cell.value
            elif cell.column == 'S':
                url = cell.value
        if not to_add:
            continue
        if args.topic and args.topic != topic:
            to_add = False
        if args.isbn and args.isbn != isbn:
            to_add = False
        if to_add:
            books.append((title, author, isbn, topic, url))
    return books

def do_list_topics(args, work_sheet):
    topics = set()
    for row in work_sheet.iter_rows():
        to_add = True
        for cell in row:
            if cell.row == 1:
                to_add = False
                break
            if cell.column == 'L':
                topic = cell.value
            elif cell.column == 'G':
                isbn = cell.value
        if not to_add:
            continue
        if args.topic and args.topic != topic:
            to_add = False
        if args.isbn and args.isbn != isbn:
            to_add = False
        if to_add:
            topics.add(topic)
    return topics

def confirm_yes_no(message):
    yes_response = {'yes', 'y', 'ye', ''}
    no_response = {'no', 'n'}
    choice = None

    while choice not in yes_response and choice not in no_response:
        sys.stdout.write(message)
        choice = input().lower()
        if choice in yes_response:
            return True
        if choice in no_response:
            return False
        sys.stdout.write("Please respond with 'yes' or 'no'.")

if __name__ == "__main__":
    run()
